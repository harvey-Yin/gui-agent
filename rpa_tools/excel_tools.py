"""
Excel数据处理工具
基于pandas和openpyxl实现Excel文件读写和数据处理
"""
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from typing import Dict, Any, List, Optional
from pathlib import Path
from datetime import datetime, date
from .base_tool import RPAToolBase


class ExcelTool(RPAToolBase):
    """Excel处理工具"""
    
    def __init__(self):
        super().__init__()
        self.description = "Excel文件读写和数据处理工具"
        self.current_workbook = None
        self.current_sheet = None
    
    def execute(self, **kwargs) -> Dict[str, Any]:
        """通用执行接口"""
        return {"status": "success", "message": "请使用具体方法"}
    
    # ========== 文件读写 ==========
    
    def read_excel(self, file_path: str, sheet_name: Optional[str] = None,
                   header: Optional[int] = 0, usecols: Optional[List] = None) -> Dict[str, Any]:
        """
        读取Excel文件
        
        Args:
            file_path: Excel文件路径
            sheet_name: 工作表名称（None=第一个sheet）
            header: 表头行号（0开始）
            usecols: 要读取的列（列表或范围）
        """
        try:
            df = pd.read_excel(
                file_path,
                sheet_name=sheet_name,
                header=header,
                usecols=usecols
            )
            
            return {
                "status": "success",
                "data": df,
                "shape": df.shape,
                "columns": list(df.columns),
                "message": f"成功读取Excel: {file_path}, 形状: {df.shape}"
            }
        except Exception as e:
            return {"status": "error", "message": str(e)}
    
    def write_excel(self, data: pd.DataFrame, file_path: str,
                   sheet_name: str = 'Sheet1', index: bool = False) -> Dict[str, Any]:
        """
        写入Excel文件
        
        Args:
            data: pandas DataFrame
            file_path: 输出文件路径
            sheet_name: 工作表名称
            index: 是否写入索引
        """
        try:
            data.to_excel(file_path, sheet_name=sheet_name, index=index)
            
            return {
                "status": "success",
                "path": file_path,
                "shape": data.shape,
                "message": f"成功写入Excel: {file_path}"
            }
        except Exception as e:
            return {"status": "error", "message": str(e)}
    
    # ========== 工作簿操作（openpyxl） ==========
    
    def open_workbook(self, file_path: str) -> Dict[str, Any]:
        """打开Excel工作簿"""
        try:
            self.current_workbook = load_workbook(file_path)
            sheet_names = self.current_workbook.sheetnames
            
            return {
                "status": "success",
                "path": file_path,
                "sheets": sheet_names,
                "message": f"打开工作簿: {file_path}, 包含 {len(sheet_names)} 个工作表"
            }
        except Exception as e:
            return {"status": "error", "message": str(e)}
    
    def select_sheet(self, sheet_name: str) -> Dict[str, Any]:
        """选择工作表"""
        try:
            if self.current_workbook is None:
                return {"status": "error", "message": "请先打开工作簿"}
            
            self.current_sheet = self.current_workbook[sheet_name]
            
            return {
                "status": "success",
                "sheet": sheet_name,
                "max_row": self.current_sheet.max_row,
                "max_column": self.current_sheet.max_column,
                "message": f"选择工作表: {sheet_name}"
            }
        except Exception as e:
            return {"status": "error", "message": str(e)}
    
    def write_cell(self, row: int, column: int, value: Any) -> Dict[str, Any]:
        """
        写入单元格
        
        Args:
            row: 行号（1开始）
            column: 列号（1开始）
            value: 要写入的值
        """
        try:
            if self.current_sheet is None:
                return {"status": "error", "message": "请先选择工作表"}
            
            self.current_sheet.cell(row=row, column=column, value=value)
            
            return {
                "status": "success",
                "row": row,
                "column": column,
                "value": value,
                "message": f"写入单元格({row},{column}): {value}"
            }
        except Exception as e:
            return {"status": "error", "message": str(e)}
    
    def read_cell(self, row: int, column: int) -> Dict[str, Any]:
        """读取单元格"""
        try:
            if self.current_sheet is None:
                return {"status": "error", "message": "请先选择工作表"}
            
            value = self.current_sheet.cell(row=row, column=column).value
            
            return {
                "status": "success",
                "row": row,
                "column": column,
                "value": value,
                "message": f"读取单元格({row},{column}): {value}"
            }
        except Exception as e:
            return {"status": "error", "message": str(e)}
    
    def save_workbook(self, file_path: Optional[str] = None) -> Dict[str, Any]:
        """保存工作簿"""
        try:
            if self.current_workbook is None:
                return {"status": "error", "message": "没有打开的工作簿"}
            
            if file_path:
                self.current_workbook.save(file_path)
            else:
                self.current_workbook.save(self.current_workbook.path)
            
            return {
                "status": "success",
                "path": file_path or self.current_workbook.path,
                "message": "工作簿已保存"
            }
        except Exception as e:
            return {"status": "error", "message": str(e)}
    
    # ========== 数据处理 ==========
    
    def filter_data(self, data: pd.DataFrame, column: str, 
                   condition: str, value: Any) -> Dict[str, Any]:
        """
        过滤数据
        
        Args:
            data: DataFrame
            column: 列名
            condition: 条件 ('==', '!=', '>', '<', '>=', '<=', 'contains')
            value: 比较值
        """
        try:
            if condition == '==':
                filtered = data[data[column] == value]
            elif condition == '!=':
                filtered = data[data[column] != value]
            elif condition == '>':
                filtered = data[data[column] > value]
            elif condition == '<':
                filtered = data[data[column] < value]
            elif condition == '>=':
                filtered = data[data[column] >= value]
            elif condition == '<=':
                filtered = data[data[column] <= value]
            elif condition == 'contains':
                filtered = data[data[column].astype(str).str.contains(str(value))]
            else:
                return {"status": "error", "message": f"不支持的条件: {condition}"}
            
            return {
                "status": "success",
                "data": filtered,
                "original_count": len(data),
                "filtered_count": len(filtered),
                "message": f"过滤完成: {len(data)} -> {len(filtered)} 行"
            }
        except Exception as e:
            return {"status": "error", "message": str(e)}
    
    def merge_data(self, left: pd.DataFrame, right: pd.DataFrame,
                  on: str, how: str = 'left') -> Dict[str, Any]:
        """
        合并数据
        
        Args:
            left: 左表
            right: 右表
            on: 连接键
            how: 连接方式 ('left', 'right', 'inner', 'outer')
        """
        try:
            merged = left.merge(right, on=on, how=how)
            
            return {
                "status": "success",
                "data": merged,
                "shape": merged.shape,
                "message": f"合并完成: {left.shape} + {right.shape} -> {merged.shape}"
            }
        except Exception as e:
            return {"status": "error", "message": str(e)}
    
    def group_aggregate(self, data: pd.DataFrame, group_by: str,
                       agg_column: str, agg_func: str) -> Dict[str, Any]:
        """
        分组聚合
        
        Args:
            data: DataFrame
            group_by: 分组列
            agg_column: 聚合列
            agg_func: 聚合函数 ('sum', 'mean', 'count', 'min', 'max')
        """
        try:
            result = data.groupby(group_by)[agg_column].agg(agg_func).reset_index()
            
            return {
                "status": "success",
                "data": result,
                "groups": len(result),
                "message": f"分组聚合完成: {len(result)} 个分组"
            }
        except Exception as e:
            return {"status": "error", "message": str(e)}
    
    # ========== 高级功能（基于现有项目） ==========
    
    def find_header_row(self, keywords: List[str], max_search_rows: int = 10) -> Dict[str, Any]:
        """
        查找表头行（基于关键字）
        
        Args:
            keywords: 必须包含的关键字列表
            max_search_rows: 最多搜索的行数
        """
        try:
            if self.current_sheet is None:
                return {"status": "error", "message": "请先选择工作表"}
            
            for row in range(1, max_search_rows + 1):
                headers = {}
                for col_idx, cell in enumerate(self.current_sheet[row], start=1):
                    if cell.value:
                        headers[cell.value] = col_idx
                
                # 检查是否包含所有关键字
                if all(kw in headers for kw in keywords):
                    return {
                        "status": "success",
                        "header_row": row,
                        "headers": headers,
                        "message": f"找到表头行: 第{row}行"
                    }
            
            return {
                "status": "error",
                "message": f"前{max_search_rows}行未找到包含所有关键字的表头"
            }
        except Exception as e:
            return {"status": "error", "message": str(e)}
    
    def write_daily_data(self, target_date: date, value: Any,
                        date_column_name: str, value_column_name: str) -> Dict[str, Any]:
        """
        写入日期数据（基于现有daily_report项目）
        
        自动查找日期匹配行，如果不存在则追加
        """
        try:
            if self.current_sheet is None:
                return {"status": "error", "message": "请先选择工作表"}
            
            # 查找表头
            header_result = self.find_header_row([date_column_name, value_column_name])
            if header_result["status"] != "success":
                return header_result
            
            header_row = header_result["header_row"]
            headers = header_result["headers"]
            date_col = headers[date_column_name]
            value_col = headers[value_column_name]
            
            # 查找匹配日期的行
            written = False
            for row_idx in range(header_row + 1, self.current_sheet.max_row + 1):
                cell_date = self.current_sheet.cell(row=row_idx, column=date_col).value
                
                # 尝试转换为date进行比较
                cmp_date = None
                if isinstance(cell_date, datetime):
                    cmp_date = cell_date.date()
                elif isinstance(cell_date, date):
                    cmp_date = cell_date
                
                if cmp_date == target_date:
                    # 找到匹配行，更新值
                    self.current_sheet.cell(row=row_idx, column=value_col, value=value)
                    written = True
                    break
            
            # 如果未找到，追加新行
            if not written:
                new_row_idx = self.current_sheet.max_row + 1
                self.current_sheet.cell(row=new_row_idx, column=date_col, value=target_date)
                self.current_sheet.cell(row=new_row_idx, column=value_col, value=value)
            
            return {
                "status": "success",
                "date": target_date,
                "value": value,
                "written": written,
                "message": f"{'更新' if written else '新增'}日期数据: {target_date} = {value}"
            }
        except Exception as e:
            return {"status": "error", "message": str(e)}
